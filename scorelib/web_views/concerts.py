"""
SKG Notenbank - Sheet Music Database and Archive Management System
Copyright (C) 2026 Arno Euteneuer

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Font

from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.utils.text import slugify

from ..models import Concert


@login_required
def concert_detail_view(request, concert_id=None):
    if concert_id:
        next_concert = get_object_or_404(Concert, pk=concert_id)
    else:
        next_concert = (
            Concert.objects.filter(date__isnull=False, date__gte=timezone.now())
            .order_by("date")
            .first()
        )
        if not next_concert:
            next_concert = (
                Concert.objects.filter(date__isnull=False).order_by("-date").first()
            )

    if not next_concert:
        return render(request, "scorelib/concert_detail.html", {"concert": None})

    context = {"concert": next_concert}
    total_duration = next_concert.programitem_set.aggregate(
        total=Sum("piece__duration")
    )["total"]

    if not total_duration:
        from datetime import timedelta

        total_duration = timedelta(0)

    context["total_duration"] = total_duration

    total_seconds = int(total_duration.total_seconds())
    minutes = total_seconds // 60
    seconds = total_seconds % 60

    if minutes > 60:
        hours = minutes // 60
        remaining_minutes = minutes % 60
        formatted_duration = f"{hours} Std. {remaining_minutes} Min."
    else:
        formatted_duration = f"{minutes}:{seconds:02d} Min."

    context["formatted_duration"] = formatted_duration

    profile = getattr(request.user, "profile", None)
    context["has_full_archive_access"] = request.user.is_staff or (
        profile.has_full_archive_access if profile else False
    )

    program_data = []
    for item in next_concert.programitem_set.all().select_related("piece"):
        piece = item.piece

        user_parts = []
        if profile and (
            profile.has_full_archive_access or piece.is_active_for_download()
        ):
            all_parts = piece.parts.all()
            user_parts = [p for p in all_parts if profile.can_view_part(p.part_name)]

        has_youtube = piece.external_links.filter(
            Q(url__icontains="youtube.com") | Q(url__icontains="youtu.be")
        ).exists()

        program_data.append(
            {
                "piece": piece,
                "user_parts": user_parts,
                "has_youtube": has_youtube,
                "recordings": piece.audiorecording_set.filter(concert=next_concert),
            }
        )

    context["program_data"] = program_data
    context["user_profile"] = profile

    return render(request, "scorelib/concert_detail.html", context)


@login_required
def concert_list_view(request):
    from django.core.paginator import Paginator

    f_search = request.GET.get("search", "")
    f_sort = request.GET.get("sort", "date")
    f_sort_dir = request.GET.get("sort_dir", "desc")

    concerts = Concert.objects.all()

    if f_search:
        concerts = concerts.filter(
            Q(title__icontains=f_search) | Q(subtitle__icontains=f_search)
        )

    if f_sort == "date":
        if f_sort_dir == "asc":
            concerts = concerts.order_by("sort_date", "title")
        else:
            concerts = concerts.order_by("-sort_date", "title")
    else:
        if f_sort_dir == "asc":
            concerts = concerts.order_by("title")
        else:
            concerts = concerts.order_by("-title")

    paginator = Paginator(concerts, 50)
    page_number = request.GET.get("page", 1)
    try:
        page_obj = paginator.page(page_number)
    except Exception:
        page_obj = paginator.page(1)

    context = {
        "concerts": page_obj.object_list,
        "page_obj": page_obj,
        "active_filters": request.GET,
        "current_sort": f_sort,
        "current_sort_dir": f_sort_dir,
        "total_count": paginator.count,
    }
    return render(request, "scorelib/concert_list.html", context)


@login_required
def export_concert_setlist_gema(request, concert_id):
    concert = get_object_or_404(Concert, pk=concert_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Setlist"

    ws.append(["Auftritt:", f"{concert.title}"])
    date_text = concert.date.strftime("%d.%m.%Y %H:%M") if concert.date else ""
    ws.append(["Datum:", f"{date_text}"])
    ws.append([])

    headers = ["Index", "Titel", "Komponist", "Arrangeur", "Verlag", "Potpourri"]
    ws.append(headers)

    bold_font = Font(bold=True)
    try:
        ws["A1"].font = bold_font
        ws["A2"].font = bold_font
    except Exception:
        pass

    header_row = ws.max_row
    for col_idx in range(1, len(headers) + 1):
        try:
            ws.cell(row=header_row, column=col_idx).font = bold_font
        except Exception:
            pass

    for idx, item in enumerate(
        concert.programitem_set.all().select_related("piece").order_by("order"), start=1
    ):
        piece = item.piece
        title = piece.title or ""
        composer = piece.composer.name if piece.composer else ""
        arranger = piece.arranger.name if piece.arranger else ""
        publisher = piece.publisher.name if piece.publisher else ""
        medley = "ja" if getattr(piece, "is_medley", False) else "nein"

        ws.append([idx, title, composer, arranger, publisher, medley])

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"setlist_{slugify(concert.title)}-{concert.id}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
