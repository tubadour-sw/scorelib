"""
SKG Notenbank - Sheet Music Database and Archive Management System
Copyright (C) 2026 Arno Euteneuer

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

import os
import csv, io
from django.core.paginator import Paginator
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.db.models import Q, Sum
from django.db import transaction
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, Http404
from django.http import FileResponse
from django.contrib.auth.models import User
from django.utils.text import slugify


# Importiere deine Modelle
from .models import Piece, Part, Concert, Arranger, Composer, Publisher, Genre, MusicianProfile, AudioRecording, InstrumentGroup
from .forms import CSVPiecesImportForm, CSVUserImportForm, UserProfileUpdateForm

@login_required
def scorelib_index(request):
    # Load recordings directly to avoid database queries in loops
    pieces = Piece.objects.select_related(
        'composer', 
        'arranger', 
        'publisher'
    ).order_by('title').prefetch_related(
        'concerts', 
        'genres',
        'audiorecording_set'
    )
    
    
    # Get filter values from URL (GET request)
    f_search = request.GET.get('search')
    f_genre = request.GET.get('genre')
    f_diff = request.GET.get('difficulty')
    f_comp = request.GET.get('composer')
    f_arr = request.GET.get('arranger')
    f_pub = request.GET.get('publisher')
    f_con = request.GET.get('concert')
    f_sort = request.GET.get('sort', 'title')  # Default sort by title
    f_sort_dir = request.GET.get('sort_dir', 'asc')  # 'asc' or 'desc'
    f_sort_artist = request.GET.get('sort_artist', 'composer')  # Choose composer or arranger for the artist column

    # Filter anwenden
    if f_search:
        pieces = pieces.filter(
            Q(title__icontains=f_search) | 
            Q(archive_label__icontains=f_search) |
            Q(composer__name__icontains=f_search) |
            Q(arranger__name__icontains=f_search) |
            Q(additional_info__icontains=f_search)
        )
    if f_genre:
        pieces = pieces.filter(genres__id=f_genre)
    if f_diff:
        pieces = pieces.filter(difficulty=f_diff)
    if f_comp:
        pieces = pieces.filter(composer__id=f_comp)
    if f_arr:
        pieces = pieces.filter(arranger__id=f_arr)
    if f_pub:
        pieces = pieces.filter(publisher__id=f_pub)
    if f_con:
        # Here we filter via the program (ProgramItems) of the selected concert
        pieces = pieces.filter(programitem__concert_id=f_con)

    # Prevent duplicates (due to ManyToMany Genres)
    pieces = pieces.distinct()

    # Apply sorting to the full queryset before pagination
    f_sort_artist = request.GET.get('sort_artist', 'composer')  # Choose composer or arranger for the artist column
    
    if f_sort == 'title':
        order_field = 'title'
    elif f_sort == 'composer':
        # Sort by chosen artist field (composer or arranger)
        if f_sort_artist == 'arranger':
            order_field = 'arranger__name'
        else:
            order_field = 'composer__name'
    elif f_sort == 'publisher':
        order_field = 'publisher__name'
    elif f_sort == 'difficulty':
        order_field = 'difficulty'
    else:
        order_field = 'title'
    
    # Apply direction prefix for descending
    if f_sort_dir == 'desc':
        order_field = f'-{order_field}'
    
    pieces = pieces.order_by(order_field)
    
    # Apply pagination: 50 items per page
    paginator = Paginator(pieces, 50)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page_number)
    except Exception:
        page_obj = paginator.page(1)
    
    context = {
        'pieces': page_obj.object_list,  # Only the pieces for this page
        'page_obj': page_obj,  # The paginator object for template
        'genres': Genre.objects.all().order_by('name'),
        'composers': Composer.objects.all().order_by('name'),
        'arrangers': Arranger.objects.all().order_by('name'),
        'publishers': Publisher.objects.all().order_by('name'),
        'concerts': Concert.objects.all().order_by('-date'),
        'active_filters': request.GET,
        'current_sort': f_sort,
        'current_sort_dir': f_sort_dir,
        'current_sort_artist': f_sort_artist,
        'total_count': paginator.count
    }
    return render(request, 'scorelib/index.html', context)

def scorelib_search(request):
    """Returns search results as JSON."""
    query = request.GET.get('q', '')
    # Search in title, composer, or archive number
    pieces = Piece.objects.filter(
        Q(title__icontains=query) | 
        Q(additional_info__icontains=query) |
        Q(composer__name__icontains=query) |
        Q(arranger__name__icontains=query) |
        Q(archive_label__icontains=query)
    ).distinct()[:20] # Limited to 20 results for speed

    results = []
    for piece in pieces:
        # Get all parts for this piece
        parts = []
        for part in piece.parts.all():
            parts.append({
                'id': part.id,
                'name': part.part_name,
                'url': part.pdf_file.url
            })
        
        results.append({
            'id': piece.id,
            'title': piece.title,
            'composer': piece.composer.name if piece.composer else '',
            'label': piece.archive_label,
            'parts': parts
        })
    
    return JsonResponse({'results': results})
    
@login_required
def concert_detail_view(request, concert_id=None):
    if concert_id:
        # Load a specific concert
        next_concert = get_object_or_404(Concert, pk=concert_id)
    else:
        # 1. Try: Find the next concert chronologically
        next_concert = Concert.objects.filter(
            date__isnull=False, 
            date__gte=timezone.now()
        ).order_by('date').first()
        
        # 2. Fallback: If no future concert exists, use the latest one
        if not next_concert:
            next_concert = Concert.objects.filter(
                date__isnull=False
            ).order_by('-date').first()
            
    # If the database is completely empty (no concert at all),
    # we should skip the rest of the view or show a message
    if not next_concert:
        return render(request, 'scorelib/concert_detail.html', {'concert': None})
        
    context = {'concert': next_concert}
    
    # Calculate the sum of all 'duration' fields of pieces in the program
    # We access the Piece via ProgramItem
    total_duration = next_concert.programitem_set.aggregate(
        total=Sum('piece__duration')
    )['total']

    # If the program is empty, set duration to 0
    if not total_duration:
        from datetime import timedelta
        total_duration = timedelta(0)
        
    context['total_duration'] = total_duration
    
    # Format duration for display
    total_seconds = int(total_duration.total_seconds())
    minutes = total_seconds // 60
    seconds = total_seconds % 60

    if minutes > 60:
        hours = minutes // 60
        remaining_minutes = minutes % 60
        formatted_duration = f"{hours} Std. {remaining_minutes} Min."
    else:
        formatted_duration = f"{minutes}:{seconds:02d} Min."

    context['formatted_duration'] = formatted_duration

    if next_concert:
        # Get the logged-in user's profile
        # (Uses the 'related_name=profile' from the model)
        profile = getattr(request.user, 'profile', None)
        # Expose a boolean for templates to check full archive access
        context['has_full_archive_access'] = request.user.is_staff or (profile.has_full_archive_access if profile else False)
        program_data = []
        # Iterate through the concert program (via ProgramItem for order)
        for item in next_concert.programitem_set.all().select_related('piece'):
            piece = item.piece
            
            # Filter parts based on user's instrument filter
            user_parts = []
            if profile and (profile.has_full_archive_access or piece.is_active_for_download()):
                all_parts = piece.parts.all()
                user_parts = [p for p in all_parts if profile.can_view_part(p.part_name)]
            
            program_data.append({
                'piece': piece,
                'user_parts': user_parts,
                'recordings': piece.audiorecording_set.filter(concert=next_concert)
            })
            
        context['program_data'] = program_data

    # Ensure templates can access the profile object if needed
    context['user_profile'] = profile

    return render(request, 'scorelib/concert_detail.html', context)

@login_required
def concert_list_view(request):
    # All concerts sorted by date (newest first)
    concerts = Concert.objects.all().order_by('title')
    return render(request, 'scorelib/concert_list.html', {'concerts': concerts})

@login_required
def protected_part_download(request, part_id):
    part = get_object_or_404(Part, pk=part_id)
    
    if not request.user.is_staff: 
        profile = getattr(request.user, 'profile', None)
        piece = part.piece
        if not profile:
            return HttpResponse("Zugriff verweigert: Du hast keinen Zugriff auf Noten.", status=403)
            
        if not profile.has_full_archive_access:
            if not piece.is_active_for_download():
                return HttpResponse("Zugriff verweigert: Noten für dieses Stück stehen momentan nicht zur Verfügung.", status=403)

            if not profile.can_view_part(part.part_name):
                return HttpResponse("Zugriff verweigert: Diese Stimme gehört nicht zu deinem Instrumenten-Filter.", status=403)

    # Pfad zur Datei auf der Festplatte
    file_path = part.pdf_file.path
    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            #response = HttpResponse(fh.read(), content_type="application/pdf")
            response = FileResponse(open(file_path, 'rb'), content_type="application/pdf") 
            # Opens the PDF in the browser instead of downloading immediately
            response['Content-Disposition'] = f'inline; filename="{os.path.basename(file_path)}"'
            return response
    raise Http404
    
@login_required
def protected_audio_download(request, audio_id):
    recording = get_object_or_404(AudioRecording, pk=audio_id)
    
    # Pfad zur Datei
    file_path = recording.audio_file.path
    
    if not os.path.exists(file_path):
        raise Http404

    # FileResponse ist effizienter fuer Streaming (Audio/Video)
    response = FileResponse(open(file_path, 'rb'), content_type="audio/mpeg")
    response['Content-Disposition'] = f'inline; filename="{os.path.basename(file_path)}"'
    return response


def piece_csv_import(request):
    if request.method == "POST":
        form = CSVPiecesImportForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['csv_file']
            
            try:
                # utf-8-sig handles Excel BOM issues
                data_set = csv_file.read().decode('utf-8-sig')
                io_string = io.StringIO(data_set)
                
                # Read the header separately for validation
                reader = csv.DictReader(io_string, delimiter=';')
                
                # Consistency check: Are required columns present?
                # reader.fieldnames contains the header names
                required_columns = ['Title', 'Composer', 'Arranger']
                missing_columns = [col for col in required_columns if col not in reader.fieldnames]
                
                if missing_columns:
                    messages.error(
                        request, 
                        f"Import abgebrochen: Die CSV-Datei hat ein falsches Format. "
                        f"Es fehlen folgende Spalten: {', '.join(missing_columns)}. "
                        f"Bitte prüfen Sie die Groß-/Kleinschreibung."
                    )
                    return redirect(request.path)
                
                created_count = 0
                updated_count = 0
                with transaction.atomic():
                    for row in reader:
                        if not row.get("Title"):
                            continue
                        
                        # Get or create composer
                        composer, _ = Composer.objects.get_or_create(name=row.get('Composer', '').strip())
                        
                        # Get or create arranger (optional)
                        arranger = None
                        if row.get('Arranger'):
                            arranger, _ = Arranger.objects.get_or_create(name=row['Arranger'].strip())
                            
                        publisher = None
                        if row.get('Publisher'):
                            publisher, _ = Publisher.objects.get_or_create(name=row['Publisher'].strip())
                        
                        # Difficulty level
                        diff_raw = row.get('Difficulty', '').strip()
                        difficulty = int(diff_raw) if diff_raw.isdigit() else None
                        
                        # Duration logic for DurationField
                        duration_raw = row.get('Duration', '').strip() # Assumes column is named this way
                        duration_delta = None
                        
                        if duration_raw and ':' in duration_raw:
                            try:
                                parts = duration_raw.split(':')
                                if len(parts) == 2: # mm:ss
                                    minutes, seconds = map(int, parts)
                                    duration_delta = timedelta(minutes=minutes, seconds=seconds)
                                elif len(parts) == 3: # hh:mm:ss
                                    hours, minutes, seconds = map(int, parts)
                                    duration_delta = timedelta(hours=hours, minutes=minutes, seconds=seconds)
                            except ValueError:
                                pass # If text is in the time column, we ignore it
                                
                        # Create piece (only if title + composer combination doesn't exist yet)
                        piece, created = Piece.objects.update_or_create(
                            title=row['Title'].strip(),
                            archive_label=row.get('Label', '').strip(),
                            composer=composer,
                            defaults={
                                'arranger': arranger,
                                'duration': duration_delta,
                                'difficulty': difficulty,
                                'publisher': publisher,
                            }
                        )
                        
                        # Link genres (multiple entries separated by comma possible)
                        if row.get('Genres'):
                            genre_names = [g.strip() for g in row['Genres'].split(',')]
                            for g_name in genre_names:
                                genre, _ = Genre.objects.get_or_create(name=g_name)
                                piece.genres.add(genre)
                        
                        if row.get('Concerts'):
                            concert_names = [g.strip() for g in row['Concerts'].split(',')]
                            for c_name in concert_names:
                                concert, _ = Concert.objects.get_or_create(title=c_name)
                                piece.concerts.add(concert)
                        
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    
                    messages.success(request, f"Import abgeschlossen: {created_count} Stücke neu angelegt, {updated_count} Stücke aktualisiert.")
                    return redirect('admin:scorelib_piece_changelist')
            except UnicodeDecodeError:
                messages.error(request, "Fehler: Die Datei konnte nicht gelesen werden. Bitte stellen Sie sicher, dass sie als CSV (UTF-8) gespeichert wurde.")
            except Exception as e:
                messages.error(request, f"Ein unerwarteter Fehler ist aufgetreten: {e}")
    else:
        form = CSVPiecesImportForm()
    
    return render(request, 'admin/csv_pieces_import.html', {'form': form})
   

def index(request):
    # Base queryset
    pieces = Piece.objects.all().select_related('composer', 'arranger', 'publisher').order_by('title')

    # Get filter values from URL (GET request)
    f_search = request.GET.get('search')
    f_genre = request.GET.get('genre')
    f_diff = request.GET.get('difficulty')
    f_comp = request.GET.get('composer')
    f_arr = request.GET.get('arranger')
    f_pub = request.GET.get('publisher')
    f_con = request.GET.get('concert')

    # Filter anwenden
    if f_search:
        pieces = pieces.filter(Q(title__icontains=f_search) | Q(archive_label__icontains=f_search))
    if f_genre:
        pieces = pieces.filter(genres__id=f_genre)
    if f_diff:
        pieces = pieces.filter(difficulty=f_diff)
    if f_comp:
        pieces = pieces.filter(composer__id=f_comp)
    if f_arr:
        pieces = pieces.filter(arranger__id=f_arr)
    if f_pub:
        pieces = pieces.filter(publisher__id=f_pub)
    if f_con:
        # Here we filter via the program (ProgramItems) of the selected concert
        pieces = pieces.filter(programitem__concert_id=f_con)

    # Dubletten verhindern (wegen ManyToMany Genres)
    pieces = pieces.distinct()

    # Data for dropdown menus in the template
    context = {
        'pieces': pieces,
        'genres': Genre.objects.all().order_by('name'),
        'composers': Composer.objects.all().order_by('name'),
        'arrangers': Arranger.objects.all().order_by('name'),
        'publishers': Publisher.objects.all().order_by('name'),
        'concerts': Concert.objects.all().order_by('-date'),
        # Return active filters to set "selected" in template
        'active_filters': request.GET
    }
    return render(request, 'scorelib/index.html', context)


from django.shortcuts import render, get_object_or_404
from .models import Piece

def piece_detail(request, pk):
    piece = get_object_or_404(Piece, pk=pk)
    user_profile = getattr(request.user, 'profile', None)
    
    # Get all parts and sort alphabetically by name
    all_parts = list(piece.parts.all())
    all_parts.sort(key=lambda x: x.part_name.lower())
    
    user_parts = []

    if request.user.is_staff or (user_profile and user_profile.has_full_archive_access):
        user_parts = all_parts
    elif user_profile and user_profile.instrument_groups.exists() and piece.is_active_for_download():
        for part in all_parts:
            if user_profile.can_view_part(part.part_name):
                user_parts.append(part)
        
        # Sort the filtered list as well for safety
        user_parts.sort(key=lambda x: x.part_name.lower())

    return render(request, 'scorelib/piece_detail.html', {
        'piece': piece,
        'user_parts': user_parts,
        'all_parts': all_parts,
        'recordings': piece.audiorecording_set.all(),
        'program_items': piece.programitem_set.select_related('concert').order_by('-concert__date'),
    })
    
def legal_view(request):
    return render(request, 'scorelib/legal.html')
    

@login_required
@transaction.atomic
def import_musicians(request):
    if not request.user.is_staff:
        return redirect('scorelib_index')

    available_groups = InstrumentGroup.objects.all().order_by('name')
    # Create a set for quick group name matching
    group_names_set = {g.name.lower(): g for g in available_groups}
        
    if request.method == "POST":
        form = CSVUserImportForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['csv_file']
            is_dry_run = form.cleaned_data.get('dry_run')
            
            try:
                data_set = csv_file.read().decode('utf-8-sig')
                io_string = io.StringIO(data_set)
                reader = csv.DictReader(io_string, delimiter=';')
            except Exception as e:
                messages.error(request, f"Datei konnte nicht gelesen werden: {e}")
                return redirect('import_musicians')

            import_results = []
            
            # Start einer globalen Transaktion
            sid = transaction.savepoint()
            
            try:
                for line_num, row in enumerate(reader, start=2):
                    try:
                        first_name = row.get("FirstName").strip()
                        last_name = row.get("LastName").strip()
                        groups_raw = row.get("Instruments").strip()
                    except:
                        import_results.append({
                            'line': line_num, 'name': "Unvollständig", 
                            'status': "Fehler: Mind. 'FirstName', 'LastName', 'Instruments' nötig", 'type': 'danger'
                        })
                        continue
                        
                    groups_final = ""
                    email_raw = row.get("Email")
                    email = email_raw.strip() if email_raw else ""
                    
                    # Generate user data
                    username = slugify(f"{first_name} {last_name}")
                    # Password: remove spaces in last name
                    raw_password = f"SKG-{last_name.replace(' ', '')}"
                    
                    # Process groups
                    target_groups = [g.strip() for g in groups_raw.split(',') if g.strip()]
                    valid_groups = []
                    unknown_groups = []
                    
                    for g_name in target_groups:
                        if g_name.lower() in group_names_set:
                            valid_groups.append(group_names_set[g_name.lower()])                                
                        else:
                            found = False
                            for group_obj in available_groups:
                                if group_obj.matches_part(g_name):
                                    valid_groups.append(group_obj)
                                    found = True
                                    break # Ersten Treffer nehmen
                            
                            if not found:
                                unknown_groups.append(g_name)
                    
                    # Inner savepoint for this row
                    row_sid = transaction.savepoint()
                    
                    try:
                        user, created = User.objects.get_or_create(
                            username=username,
                            defaults={
                                'first_name': first_name,
                                'last_name': last_name,
                                'email': email
                            }
                        )
                        
                        if created:
                            user.set_password(raw_password)
                            user.save()
                            status_text = "Neu angelegt"
                            row_type = "success"
                        else:
                            status_text = "Bereits vorhanden (aktualisiert)"
                            raw_password = "(unverändert)"
                            row_type = "warning"

                        # Profil & Instrumentengruppen
                        profile, _ = MusicianProfile.objects.get_or_create(user=user)
                        
                        
                        if valid_groups:
                            profile.instrument_groups.set(valid_groups)
                            groups_final = ", ".join(g.name for g in valid_groups)
                        
                        if unknown_groups:
                            status_text += f" | Unbekannte Gruppen: {', '.join(unknown_groups)}"
                            row_type = "warning"

                        transaction.savepoint_commit(row_sid)
                        
                        
                    except Exception as e:
                        transaction.savepoint_rollback(row_sid)
                        status_text = f"Kritischer Fehler: {str(e)}"
                        row_type = "danger"
                        raw_password = "-"

                    import_results.append({
                        'line': line_num,
                        'name': f"{first_name} {last_name}",
                        'email': email,
                        'username': username,
                        'password': raw_password,
                        'instrument_groups': groups_final,
                        'status': status_text,
                        'type': row_type
                    })

                # Abschluss der Transaktion
                if is_dry_run:
                    transaction.savepoint_rollback(sid)
                else:
                    transaction.savepoint_commit(sid)

                return render(request, "admin/csv_user_import_results.html", {
                    "results": import_results,
                    "is_dry_run": is_dry_run,
                    "title": "Import Ergebnis"
                })

            except Exception as e:
                transaction.savepoint_rollback(sid)
                messages.error(request, f"Allgemeiner Fehler beim Import: {e}")
                return redirect('import_musicians')
    else:
        form = CSVUserImportForm()
    
    return render(request, "admin/csv_user_import.html", {
        "form": form, 
        "title": "Musiker-Import",
        "available_groups": available_groups
    })


@login_required
def export_import_results_csv(request):
    if not request.user.is_staff:
        return redirect('scorelib_index')

    # Get data from POST request (hidden fields in form)
    # or alternatively: generate from just-processed data.
    # Since we need them right after import, a download button
    # on the results page makes the most sense.
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="musiker_zugangsdaten.csv"'
    
    writer = csv.writer(response, delimiter=';')
    writer.writerow(['Name', 'Email', 'Username', 'InitialPassword', 'InstrumentGroups', 'Status'])

    # Die Daten werden per POST vom Ergebnis-Template gesendet
    names = request.POST.getlist('name[]')
    emails = request.POST.getlist('email[]')
    usernames = request.POST.getlist('username[]')
    passwords = request.POST.getlist('password[]')
    instrument_groups = request.POST.getlist('instrument_groups[]')
    statuses = request.POST.getlist('status[]')

    for n, e, u, p, i, s in zip(names, emails, usernames, passwords, instrument_groups, statuses):
        writer.writerow([n, e, u, p, i, s])

    return response


@login_required
def export_concert_setlist_gema(request, concert_id):
    """Export the program of a concert as an Excel (.xlsx) file.

    Required columns:
    1. index (within the concert)
    2. Title
    3. Composer
    4. Arranger
    5. Publisher
    6. Medley (yes/no)

    Also write a small header containing concert name and date.
    """
    concert = get_object_or_404(Concert, pk=concert_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Setlist"

    # Header rows with concert info
    ws.append(["Auftritt:", f"{concert.title}"])
    date_text = concert.date.strftime('%d.%m.%Y %H:%M') if concert.date else ''
    ws.append(["Datum:", f"{date_text}"])
    ws.append([])

    # Column headers
    headers = ['Index', 'Titel', 'Komponist', 'Arrangeur', 'Verlag', 'Potpourri']
    ws.append(headers)

    # Bold the label cells and the header row
    bold_font = Font(bold=True)
    try:
        ws['A1'].font = bold_font
        ws['A2'].font = bold_font
    except Exception:
        pass

    header_row = ws.max_row
    for col_idx in range(1, len(headers) + 1):
        try:
            ws.cell(row=header_row, column=col_idx).font = bold_font
        except Exception:
            pass

    # Fill rows
    for idx, item in enumerate(concert.programitem_set.all().select_related('piece').order_by('order'), start=1):
        piece = item.piece
        title = piece.title or ''
        composer = piece.composer.name if piece.composer else ''
        arranger = piece.arranger.name if piece.arranger else ''
        publisher = piece.publisher.name if piece.publisher else ''
        medley = 'ja' if getattr(piece, 'is_medley', False) else 'nein'

        ws.append([idx, title, composer, arranger, publisher, medley])

    # Autosize some columns (basic)
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Write workbook to in-memory bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"setlist_{slugify(concert.title)}-{concert.id}.xlsx"
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

@login_required
def profile_view(request):
    user_profile = getattr(request.user, 'profile', None)
    
    if request.method == 'POST':
        form = UserProfileUpdateForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Profil erfolgreich aktualisiert!")
            return redirect('profile_view')
    else:
        form = UserProfileUpdateForm(instance=request.user)
    
    context = {
        'form': form,
        'user_profile': user_profile,
    }
    return render(request, 'registration/profile.html', context)